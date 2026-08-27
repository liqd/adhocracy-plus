import io

import openpyxl
import pytest

from adhocracy4.dashboard import components
from adhocracy4.test.helpers import setup_phase
from apps.customfields.models import CustomFieldAnswer
from apps.customfields.models import CustomFieldType
from apps.ideas import phases

export_component = components.modules.get("idea_export")


@pytest.mark.django_db
def test_idea_export_contains_custom_field_columns(
    client,
    phase_factory,
    idea_factory,
    category_factory,
    custom_field_factory,
    custom_field_choice_factory,
):
    phase, module, project, idea = setup_phase(
        phase_factory, idea_factory, phases.IssuePhase
    )
    module.blueprint_type = "BS"
    module.save()
    from apps.customfields.models import CustomFieldSettings

    settings = CustomFieldSettings.objects.create(module=module)
    open_field = custom_field_factory(
        settings=settings, type=CustomFieldType.OPEN, required=False
    )
    choice_field = custom_field_factory(
        settings=settings, type=CustomFieldType.CHOICE, required=False
    )
    choice_1 = custom_field_choice_factory(field=choice_field, label="Option 1")
    CustomFieldAnswer.objects.create(idea=idea, field=open_field, value="Kreuzberg")
    CustomFieldAnswer.objects.create(
        idea=idea, field=choice_field, value=str(choice_1.pk)
    )

    initiator = module.project.organisation.initiators.first()
    url = export_component.get_base_url(module)
    client.login(username=initiator.email, password="password")
    export_url = client.get(url).context["export"]
    response = client.get(export_url)
    assert response.status_code == 200

    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    header = [cell.value for cell in sheet[1]]
    assert open_field.label in header
    assert choice_field.label in header

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    row = rows[0]
    assert row[header.index(open_field.label)] == "Kreuzberg"
    assert row[header.index(choice_field.label)] == "Option 1"


@pytest.mark.django_db
def test_idea_export_without_custom_fields_has_no_extra_columns(
    client, phase_factory, idea_factory
):
    phase, module, project, idea = setup_phase(
        phase_factory, idea_factory, phases.IssuePhase
    )
    module.blueprint_type = "BS"
    module.save()
    initiator = module.project.organisation.initiators.first()
    url = export_component.get_base_url(module)
    client.login(username=initiator.email, password="password")
    export_url = client.get(url).context["export"]
    response = client.get(export_url)
    assert response.status_code == 200

    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    header = [cell.value for cell in workbook.active[1]]
    assert not any("custom_field" in str(name) for name in header)
