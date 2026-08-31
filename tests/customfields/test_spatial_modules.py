import io

import openpyxl
import pytest

from adhocracy4.dashboard import components
from adhocracy4.test.helpers import freeze_phase
from adhocracy4.test.helpers import redirect_target
from adhocracy4.test.helpers import setup_phase
from apps.customfields.models import CustomFieldAnswer
from apps.customfields.models import CustomFieldSettings
from apps.customfields.models import CustomFieldType
from apps.mapideas import models as mapidea_models
from apps.mapideas import phases as mapidea_phases

export_component = components.modules.get("mapidea_export")


@pytest.fixture
def mbs_module(phase_factory, area_settings_factory):
    """Module of the spatial brainstorming blueprint with map + custom field settings."""
    phase, module, project, _ = setup_phase(
        phase_factory, None, mapidea_phases.CollectPhase
    )
    module.blueprint_type = "MBS"
    module.save()
    area_settings_factory(module=module)
    CustomFieldSettings.objects.create(module=module)
    return module


@pytest.mark.django_db
def test_mapidea_create_saves_custom_fields(
    client,
    mbs_module,
    user,
    category_factory,
    custom_field_factory,
    custom_field_choice_factory,
):
    module = mbs_module
    phase = module.phase_set.first()
    settings = module.customfieldsettings_settings
    open_field = custom_field_factory(settings=settings, type=CustomFieldType.OPEN)
    choice_field = custom_field_factory(
        settings=settings, type=CustomFieldType.CHOICE, required=False
    )
    choice_1 = custom_field_choice_factory(field=choice_field, label="18-30")
    category = category_factory(module=module)

    with freeze_phase(phase):
        client.login(username=user.email, password="password")
        # get the url via the map idea create view
        from django.urls import reverse

        url = reverse(
            "a4_candy_mapideas:mapidea-create",
            kwargs={
                "organisation_slug": module.project.organisation.slug,
                "module_slug": module.slug,
            },
        )
        data = {
            "name": "Map Idea",
            "description": "description",
            "category": category.pk,
            "organisation_terms_of_use": True,
            "point": '{"type":"Feature","properties":{},"geometry":{"type":"Point","coordinates":[13.44,52.52]}}',
            "point_label": "Test location",
            "custom_field_{}".format(open_field.pk): "Kreuzberg",
            "custom_field_{}".format(choice_field.pk): str(choice_1.pk),
        }
        response = client.post(url, data)
        assert redirect_target(response) == "mapidea-detail"

    created = mapidea_models.MapIdea.objects.get(name="Map Idea")
    answers = created.custom_field_answers.all()
    assert answers.count() == 2
    assert answers.get(field=open_field).value == "Kreuzberg"
    assert answers.get(field=choice_field).value == str(choice_1.pk)


@pytest.mark.django_db
def test_mapidea_detail_shows_custom_field_answers(
    client,
    mbs_module,
    map_idea_factory,
    custom_field_factory,
    custom_field_choice_factory,
):
    module = mbs_module
    settings = module.customfieldsettings_settings
    open_field = custom_field_factory(settings=settings, type=CustomFieldType.OPEN)
    choice_field = custom_field_factory(
        settings=settings, type=CustomFieldType.CHOICE, required=False
    )
    choice_1 = custom_field_choice_factory(field=choice_field, label="Option 1")

    mapidea = map_idea_factory(module=module)
    CustomFieldAnswer.objects.create(
        content_object=mapidea, field=open_field, value="Kreuzberg"
    )
    CustomFieldAnswer.objects.create(
        content_object=mapidea, field=choice_field, value=str(choice_1.pk)
    )

    response = client.get(mapidea.get_absolute_url())
    assert response.status_code == 200
    content = response.content.decode()
    assert open_field.label in content
    assert "Kreuzberg" in content
    assert choice_field.label in content
    assert "Option 1" in content


@pytest.mark.django_db
def test_mapidea_export_contains_custom_field_columns(
    client,
    mbs_module,
    map_idea_factory,
    custom_field_factory,
    custom_field_choice_factory,
):
    module = mbs_module
    settings = module.customfieldsettings_settings
    open_field = custom_field_factory(settings=settings, type=CustomFieldType.OPEN)
    choice_field = custom_field_factory(
        settings=settings, type=CustomFieldType.CHOICE, required=False
    )
    choice_1 = custom_field_choice_factory(field=choice_field, label="Option 1")

    mapidea = map_idea_factory(module=module)
    CustomFieldAnswer.objects.create(
        content_object=mapidea, field=open_field, value="Kreuzberg"
    )
    CustomFieldAnswer.objects.create(
        content_object=mapidea, field=choice_field, value=str(choice_1.pk)
    )

    initiator = module.project.organisation.initiators.first()
    url = export_component.get_base_url(module)
    client.login(username=initiator.email, password="password")
    export_url = client.get(url).context["export"]
    response = client.get(export_url)
    assert response.status_code == 200

    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    sheet = workbook.active
    header = [cell.value for cell in sheet[1]]
    assert open_field.label in header
    assert choice_field.label in header

    rows = list(sheet.iter_rows(min_row=2, values_only=True))
    assert len(rows) == 1
    row = rows[0]
    assert row[header.index(open_field.label)] == "Kreuzberg"
    assert row[header.index(choice_field.label)] == "Option 1"


@pytest.mark.django_db
def test_mapidea_export_without_custom_fields_has_no_extra_columns(
    client, mbs_module, map_idea_factory
):
    module = mbs_module
    map_idea_factory(module=module)
    initiator = module.project.organisation.initiators.first()
    url = export_component.get_base_url(module)
    client.login(username=initiator.email, password="password")
    export_url = client.get(url).context["export"]
    response = client.get(export_url)
    assert response.status_code == 200

    workbook = openpyxl.load_workbook(io.BytesIO(response.content))
    header = [cell.value for cell in workbook.active[1]]
    assert not any("custom_field" in str(name) for name in header)
